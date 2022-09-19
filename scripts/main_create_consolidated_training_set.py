#from msilib.schema import Directory [This line is commented out, may need to be commented in for windows machine but must be commented out for mac]
import pandas as pd
import os
import glob
from openpyxl import load_workbook

if __name__=="__main__":

    #Match results Directory
    match_results_dir = "../output/match_results/"
    output_file = os.path.join(match_results_dir, "consolidated_file.csv")

    #Get list of all files in the match_results directory
    match_results_list = sorted([x for x in glob.glob(match_results_dir + "*.xlsx")])

    #Check if each Excel file has a Matched_blocks tab

    df_list = []
    counter = 0
    flag = False
    for excel_file in match_results_list:
        wb = load_workbook(excel_file, read_only=True)   # open an Excel file and return a workbook
        
        if 'Matched_blocks' in wb.sheetnames:
            df_tmp = pd.read_excel(excel_file, sheet_name="Matched_blocks")
            df_list = df_list + [df_tmp[["subtitle_text","audio_text"]]]
            counter = counter + 1
            if (counter == 10):
                df = pd.concat(df_list)
                df.to_csv(output_file,index=False,encoding='utf-8', mode='a',header=not os.path.exists(output_file))
                print("saved 10 files to csv...")
                counter = 0
                flag = True
                df_list = []

    if (flag == False):
        df = pd.concat(df_list)
        df.to_csv(output_file,index=False,encoding='utf-8', mode='a',header=not os.path.exists(output_file))
        print("saved %s files to csv " % len(df_list))


